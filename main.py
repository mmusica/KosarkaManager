from openpyxl import Workbook, load_workbook
from student import Student, ispis_studenata, evidencija, nadi_studenta
from openpyxl.utils import get_column_letter, get_column_interval, column_index_from_string

wb = load_workbook('documents/košarka_evidencija.xlsx')
ws = wb.active

odabir = 0
while True:
    print("1. za Unos dolazaka")
    print("2. za Pronalazak studenta")
    print("0. za izlaz")
    odabir = int(input("Vas odabir?"))
    studenti = []
    if odabir == 0:
        break
    elif odabir == 1:
        row_number = 3
        evidencija(studenti=studenti, row_number=row_number, ws=ws)
        wb.save('documents/košarka_evidencija.xlsx')

    elif odabir == 2:
        ime_studenta = input("Ime studenta kojeg trazite: ")
        pronadeni_student = nadi_studenta(ime_studenta,ws,row_number=3)
        if pronadeni_student[0]:
            print("Ime:", pronadeni_student[1].ime, "Dolasci:", pronadeni_student[1].broj_dolazaka,
                  "Izostanci:", pronadeni_student[1].broj_izostanaka)

            print("Zelite ovom studentu dodati izostanak ili dolazak?: ")
            dodaj = input("1 za dodaj, 0 za ne: ")
            if(dodaj == "1"):
                while True:

                    p_m = input(("+/-: "))
                    if p_m == "+" or p_m == "-":
                        break

                ws.cell(pronadeni_student[2],pronadeni_student[3]).value = p_m
                wb.save('documents/košarka_evidencija.xlsx')
                print("Spremljeno!")

            else:
                print("")
        else:
            print(pronadeni_student[1])

    else:
      print("Ponovno unesite vrijednost 1 ili 2 ili 0")





