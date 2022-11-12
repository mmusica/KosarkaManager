from openpyxl.utils import column_index_from_string
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

class Student():

    def __init__(self, ime_studenta):
        self.ime = ime_studenta
        self.broj_izostanaka = 0
        self.broj_dolazaka = 0

    def ponasanje(self):
        if self.broj_izostanaka > 3:
            return "BAD"
        else:
            return "GOOD"

   # def broji_izostanak(self,broj_izostanaka):
       # self.broj_izostanaka = broj_izostanaka

    #def broji_dolazak(selfs,broj_dolazaka):
       # self.broj_dolazaka = broj_dolazaka

def ispis_studenata(studenti):
    for student in studenti:
        print(student.ime, "- Broj dolazaka:", student.broj_dolazaka, "- Broj izostanaka:", student.broj_izostanaka,
              "- Ponasanje:", student.ponasanje())

def evidencija(row_number, ws, trazeni_student=None, studenti=[],evidentiraj = True):

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')

    greenFill = PatternFill(start_color='059033',
                          end_color='059033',
                          fill_type='solid')

    for row in ws.iter_rows(3, 31, column_index_from_string('A'), column_index_from_string('U'), True):

        col_number = column_index_from_string('A')

        for cell in row:
            if cell != None and cell != "+" and cell != '-':
                osoba = Student(cell)
                studenti.append(osoba)

            if cell == "+":
                osoba.broj_dolazaka += 1
                ws.cell(row_number, col_number).fill = greenFill

            if cell == None and not evidentiraj:
                pamti_column = col_number
                break
            if cell == None and evidentiraj:
                while True:
                    dolazak = str(input(("Dolazak za {} (+/-): ".format(osoba.ime))))

                    if dolazak == "+" or dolazak == "-":
                        break
                    else:
                        print("Molimo Vas da unesete + ili -")
                ws.cell(row_number, col_number).value = dolazak
                if dolazak == "+":
                    ws.cell(row_number, col_number).fill = greenFill
                else:
                    ws.cell(row_number, col_number).fill = redFill
                break

            if cell == '-':
                osoba.broj_izostanaka += 1
                ws.cell(row_number, col_number).fill = redFill

            col_number += 1

        if evidentiraj==False:
            if osoba.ime == trazeni_student:
                return [True, osoba, row_number, pamti_column]

        row_number += 1

    if evidentiraj==False:
        return [False, "Ovakav student ne postoji"]

def nadi_studenta(ime_studenta,ws,row_number):

    return evidencija(row_number=row_number, ws=ws, trazeni_student=ime_studenta, evidentiraj=False)
